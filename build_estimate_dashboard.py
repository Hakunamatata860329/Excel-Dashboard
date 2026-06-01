import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

INPUT_DIR  = Path(__file__).parent / "data" / "output"   # ESTIMATE_with.xlsx lives in output/
OUTPUT_DIR = Path(__file__).parent / "data" / "output"

C_DARK_GREEN    = "0F4C1F"
C_MID_GREEN     = "116A1E"
C_LIGHT_GREEN   = "A6D260"
C_ESTIMATE_BLUE = "9DC3E6"   # light blue for Estimate series
C_WHITE         = "FFFFFF"
C_BORDER        = "D0E2CB"

SPRINTS = [
    "Designer Rev1.14.0 Plan",
    "Designer Rev1.14.0 Sprint1",
    "Designer Rev1.14.0 Sprint2",
    "Designer Rev1.14.0 Sprint3",
    "Rev1.14.0 DIADesigner",
]

# short sprint labels for chart axis
SPRINT_LABELS = ["Plan", "Sprint1", "Sprint2", "Sprint3", "DIADesigner"]


# ── time parsing ──────────────────────────────────────────────────────────────
def parse_hours(text) -> float:
    if pd.isna(text):
        return 0.0
    text = str(text).strip()
    h = re.search(r"(\d+)\s*hour", text)
    m = re.search(r"(\d+)\s*min", text)
    return round((int(h.group(1)) if h else 0) + (int(m.group(1)) if m else 0) / 60, 1)


# ── data loading ──────────────────────────────────────────────────────────────
def load_data() -> pd.DataFrame:
    df = pd.read_excel(INPUT_DIR / "ESTIMATE_Task.xlsx")
    df["est_h"]   = df["Estimate"].apply(parse_hours)
    df["spend_h"] = df["TimeSpend"].apply(parse_hours)
    df["diff_h"]  = (df["spend_h"] - df["est_h"]).round(1)
    df["Planned For"] = df["Planned For"].str.strip()
    return df


# ── style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _apply_header_row(ws, row, texts):
    for col, val in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(C_DARK_GREEN)
        c.font = _font(bold=True, size=11)
        c.alignment = _center()
        c.border = _border()

def _kpi_cell(ws, row, col, label, value, unit=""):
    lc = ws.cell(row=row,   column=col, value=label)
    vc = ws.cell(row=row+1, column=col, value=f"{value}{unit}")
    lc.fill = _fill(C_MID_GREEN)
    lc.font = _font(bold=True, size=10)
    lc.alignment = _center()
    vc.fill = _fill("EAF4E6")
    vc.font = Font(bold=True, color="0F4C1F", size=14, name="Arial")
    vc.alignment = _center()
    for c in (lc, vc):
        c.border = _border()


# ── chart helper sheet ────────────────────────────────────────────────────────
def _make_chart_sheet(wb: Workbook) -> object:
    ws = wb.create_sheet("_ChartData")
    ws.sheet_state = "hidden"
    return ws

def _write_owner_pivot(cws, df: pd.DataFrame, owners, start_row: int, start_col: int):
    """Write owner-level pivot to chart sheet. Returns (header_row, last_data_row)."""
    cws.cell(row=start_row, column=start_col,   value="Owner")
    cws.cell(row=start_row, column=start_col+1, value="Estimate(h)")
    cws.cell(row=start_row, column=start_col+2, value="TimeSpend(h)")
    r = start_row + 1
    for owner in owners:
        sub = df[df["Owner"] == owner]
        cws.cell(row=r, column=start_col,   value=owner.split(".")[0])
        cws.cell(row=r, column=start_col+1, value=round(sub["est_h"].sum(),   1))
        cws.cell(row=r, column=start_col+2, value=round(sub["spend_h"].sum(), 1))
        r += 1
    return start_row, r - 1

def _write_sprint_pivot(cws, df: pd.DataFrame, start_row: int, start_col: int):
    """Write sprint-level pivot to chart sheet. Returns (header_row, last_data_row)."""
    cws.cell(row=start_row, column=start_col,   value="Sprint")
    cws.cell(row=start_row, column=start_col+1, value="Estimate(h)")
    cws.cell(row=start_row, column=start_col+2, value="TimeSpend(h)")
    r = start_row + 1
    for label, sprint in zip(SPRINT_LABELS, SPRINTS):
        sub = df[df["Planned For"] == sprint]
        est   = round(sub["est_h"].sum(),   1)
        spend = round(sub["spend_h"].sum(), 1)
        if est == 0 and spend == 0:
            continue   # skip sprints with no data
        cws.cell(row=r, column=start_col,   value=label)
        cws.cell(row=r, column=start_col+1, value=est)
        cws.cell(row=r, column=start_col+2, value=spend)
        r += 1
    return start_row, r - 1


# ── bar chart factory ─────────────────────────────────────────────────────────
def _bar_chart(cws, title, hdr_row, last_row, cat_col, val_cols, width, height, y_title=""):
    chart = BarChart()
    chart.type      = "bar"        # horizontal: categories on Y, values on X
    chart.grouping  = "clustered"
    chart.title     = title
    # openpyxl horizontal bar: x_axis = category axis (left/vertical),
    #                           y_axis = value axis (bottom/horizontal)
    chart.x_axis.title = y_title   # category label on left axis
    chart.y_axis.title = "Hours"   # value label on bottom axis
    chart.width     = width
    chart.height    = height

    cats = Reference(cws, min_col=cat_col, min_row=hdr_row+1, max_row=last_row)
    colors = [C_ESTIMATE_BLUE, C_LIGHT_GREEN]   # Estimate=light blue, TimeSpend=green

    for i, col in enumerate(val_cols):
        data = Reference(cws, min_col=col, min_row=hdr_row, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        s = chart.series[i]
        s.graphicalProperties.solidFill = colors[i]
        s.graphicalProperties.line.solidFill = colors[i]

    chart.set_categories(cats)
    return chart


# ── SUMMARY sheet ─────────────────────────────────────────────────────────────
def build_summary(wb: Workbook, cws, df: pd.DataFrame, owners):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(1, 50):
        ws.row_dimensions[row].height = 22

    # title banner
    ws.merge_cells("A1:N2")
    t = ws["A1"]
    t.value     = "ESTIMATE DASHBOARD  |  Owner Summary"
    t.fill      = _fill(C_DARK_GREEN)
    t.font      = Font(bold=True, color=C_WHITE, size=20, name="Arial")
    t.alignment = _center()

    # KPI row
    total_est   = round(df["est_h"].sum(),   1)
    total_spend = round(df["spend_h"].sum(), 1)
    util        = f"{total_spend/total_est*100:.1f}%" if total_est else "—"
    for (label, val, unit), col in zip(
        [("Total Estimate",  total_est,   "h"),
         ("Total TimeSpend", total_spend, "h"),
         ("Utilization Rate", util,       ""),
         ("Total Tasks",     len(df),     "")],
        [1, 4, 7, 10]
    ):
        _kpi_cell(ws, 4, col, label, val, unit)

    # write pivot data to chart sheet (rows 1–15 of _ChartData)
    owner_hdr, owner_last = _write_owner_pivot(cws, df, owners, 1, 1)
    sprint_hdr, sprint_last = _write_sprint_pivot(cws, df, 1, 5)

    # Owner chart — right side, col O, row 1
    chart1 = _bar_chart(cws, "Estimate vs TimeSpend by Owner",
                        owner_hdr, owner_last, 1, [2, 3], 24, 18, y_title="Owner")
    ws.add_chart(chart1, "O1")

    # Sprint chart — right side, below owner chart (~row 28)
    chart2 = _bar_chart(cws, "Estimate vs TimeSpend by Sprint",
                        sprint_hdr, sprint_last, 5, [6, 7], 24, 14, y_title="Planned For")
    ws.add_chart(chart2, "O28")


# ── OWNER sheet ───────────────────────────────────────────────────────────────
def build_owner(wb: Workbook, cws, owner: str, df_owner: pd.DataFrame, cws_row: int):
    safe_name = owner.split(".")[0][:28]
    ws = wb.create_sheet(safe_name)
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([10, 55, 28, 10, 18, 13, 13, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in range(1, 80):
        ws.row_dimensions[row].height = 20

    # header
    ws.merge_cells("A1:H2")
    h = ws["A1"]
    h.value     = owner
    h.fill      = _fill(C_DARK_GREEN)
    h.font      = Font(bold=True, color=C_WHITE, size=16, name="Arial")
    h.alignment = _center()

    # KPI row
    total_est   = round(df_owner["est_h"].sum(),   1)
    total_spend = round(df_owner["spend_h"].sum(), 1)
    util        = f"{total_spend/total_est*100:.1f}%" if total_est else "—"
    for (label, val, unit), col in zip(
        [("Estimate(h)",  total_est,   "h"),
         ("TimeSpend(h)", total_spend, "h"),
         ("Utilization",  util,        ""),
         ("Tasks",        len(df_owner), "")],
        [2, 3, 4, 5]
    ):
        _kpi_cell(ws, 4, col, label, val, unit)

    # write sprint pivot to chart sheet
    hdr_r, last_r = _write_sprint_pivot(cws, df_owner, cws_row, 9)

    # chart — right side of data (col A-H), placed at col J
    chart = _bar_chart(cws, "Estimate vs TimeSpend by Sprint",
                       hdr_r, last_r, 9, [10, 11], 24, 16, y_title="Planned For")
    ws.add_chart(chart, "J1")

    # detail table
    TABLE_HEADERS = ["ID", "Summary", "Planned For", "Type", "State",
                     "Estimate(h)", "TimeSpend(h)", "Diff(h)"]
    TABLE_START = 7
    _apply_header_row(ws, TABLE_START, TABLE_HEADERS)
    ws.row_dimensions[TABLE_START].height = 24

    red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for r_offset, (_, row) in enumerate(df_owner.iterrows(), 1):
        r = TABLE_START + r_offset
        for col, val in enumerate(
            [row["ID"], row["Summary"], row["Planned For"],
             row["Type"], row["State"],
             row["est_h"], row["spend_h"], row["diff_h"]], 1
        ):
            c = ws.cell(row=r, column=col, value=val)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 2))
            c.fill = _fill("FFFFFF" if r_offset % 2 == 1 else "F2F8F0")
            c.font = Font(name="Arial", size=10, color="222222")

    # conditional format on Diff col H (entire column range at once)
    last_table_row = TABLE_START + len(df_owner)
    diff_range = f"H{TABLE_START+1}:H{last_table_row}"
    ws.conditional_formatting.add(
        diff_range, CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill)
    )
    ws.conditional_formatting.add(
        diff_range, CellIsRule(operator="lessThan",    formula=["0"], fill=green_fill)
    )

    return last_r + 2   # next available row in _ChartData


# ── entry point ───────────────────────────────────────────────────────────────
def run():
    df     = load_data()
    wb     = Workbook()
    wb.remove(wb.active)
    cws    = _make_chart_sheet(wb)   # hidden helper sheet for all chart data
    owners = sorted(df["Owner"].unique())

    build_summary(wb, cws, df, owners)

    # owner pivot data for summary ends at row ~13; owner sprint pivots start at row 20
    cws_row = 20
    for owner in owners:
        cws_row = build_owner(wb, cws, owner, df[df["Owner"] == owner].copy(), cws_row)

    out = OUTPUT_DIR / "ESTIMATE_dashboard.xlsx"
    wb.save(out)
    print(f"Saved: {out}  ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    run()
