import sys
import argparse
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FCL_PATH = Path("data/TestCase/OneSW-Form-0023-TC_DIADesigner Function Check List.xlsx")
FCL_SHEET = "DIADesigner Function List"
TC_PATH = Path("data/TestCase/Test Case.csv")
OUTPUT_DIR = Path("data/output")

# Col indices (0-based) in FCL sheet
COL_L1 = 0   # 功能分類
COL_L2 = 1   # 功能項目
COL_L3 = 2   # 功能子項展開
COL_TAG = 3  # 功能關鍵字 - Tag
COL_STATUS = 6  # 產品功能 (W3A 支援狀態)

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
YES_FILL = PatternFill("solid", fgColor="E2EFDA")
NO_FILL = PatternFill("solid", fgColor="FCE4D6")
TBC_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _normalize_status(raw) -> str:
    if raw is None:
        return "TBC"
    s = str(raw).strip()
    if s.startswith("●"):
        return "●"
    if s.startswith("○"):
        return "○"
    if s.upper() == "TBC":
        return "TBC"
    return "TBC"


def parse_fcl() -> list[dict]:
    wb = openpyxl.load_workbook(FCL_PATH, read_only=True, data_only=True)
    ws = wb[FCL_SHEET]

    records = []
    current_l1 = ""
    current_l2 = ""

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header row

        l1 = row[COL_L1]
        l2 = row[COL_L2]
        l3 = row[COL_L3]
        tag = row[COL_TAG]
        status_raw = row[COL_STATUS] if len(row) > COL_STATUS else None

        if l1 is not None:
            current_l1 = str(l1).strip()
        if l2 is not None:
            current_l2 = str(l2).strip()

        # Skip rows without a usable Tag
        if tag is None:
            continue
        tag_str = str(tag).strip()
        if not tag_str or tag_str == "N/A" or "\n" in tag_str:
            continue

        # Skip footer
        if "Tag Count" in current_l2:
            continue

        l3_str = str(l3).strip() if l3 else ""
        if l3_str == "N/A":
            l3_str = ""

        status = _normalize_status(status_raw)
        included = "是" if status == "●" else ("否" if status == "○" else "TBC")

        records.append({
            "功能分類": current_l1,
            "Test Suite（功能項目）": current_l2,
            "功能子項": l3_str,
            "功能Tag": tag_str,
            "W3A 支援狀態": status,
            "納入 Suite": included,
        })

    return records


def write_sheet1(ws, records: list[dict]):
    headers = ["功能分類", "Test Suite（功能項目）", "功能子項", "功能Tag", "W3A 支援狀態", "納入 Suite"]
    col_widths = [18, 28, 16, 40, 14, 12]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    status_fill = {"●": YES_FILL, "○": NO_FILL, "TBC": TBC_FILL}

    for r, rec in enumerate(records, 2):
        for c, key in enumerate(headers, 1):
            cell = ws.cell(r, c, rec[key])
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = BORDER
        # Colour row by status
        fill = status_fill.get(rec["W3A 支援狀態"], TBC_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).fill = fill


def write_sheet2(ws, records: list[dict]):
    supported = [r for r in records if r["納入 Suite"] == "是"]

    headers = ["功能分類", "Test Suite（功能項目）", "功能子項", "功能Tag", "對應 Test Case IDs", "備註"]
    col_widths = [18, 28, 16, 40, 35, 20]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for r, rec in enumerate(supported, 2):
        ws.cell(r, 1, rec["功能分類"]).border = BORDER
        ws.cell(r, 2, rec["Test Suite（功能項目）"]).border = BORDER
        ws.cell(r, 3, rec["功能子項"]).border = BORDER
        ws.cell(r, 4, rec["功能Tag"]).border = BORDER
        ws.cell(r, 5, "").border = BORDER  # TE 填入
        ws.cell(r, 6, "").border = BORDER  # 備註
        for c in range(1, 7):
            ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, c).fill = YES_FILL

    note = ws.cell(len(supported) + 3, 1, "※ 請 TE 在「對應 Test Case IDs」欄位填入 Test Case.csv 的 Name 欄（如 1.101.9.21.2），多個以逗號分隔")
    note.font = Font(italic=True, color="888888")


def write_sheet3(ws):
    df = pd.read_csv(TC_PATH, encoding="utf-8-sig")
    keep_cols = ["項目", "子項目", "功能項", "Description", "Name", "State", "腳本狀態", "JOB分類"]
    df = df[[c for c in keep_cols if c in df.columns]].copy()
    df = df.sort_values(["子項目", "功能項"]).reset_index(drop=True)

    headers = list(df.columns)
    col_widths = {"項目": 20, "子項目": 22, "功能項": 28, "Description": 50,
                  "Name": 18, "State": 12, "腳本狀態": 10, "JOB分類": 12}

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 15)

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for r, row_data in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c, val if pd.notna(val) else "")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = BORDER


# ── Sheet 4 helpers ──────────────────────────────────────────────────────────

L1_FILL = PatternFill("solid", fgColor="2F5496")   # 深藍 - 功能分類
L1_FONT = Font(bold=True, color="FFFFFF", size=11)
L2_FILL = PatternFill("solid", fgColor="9DC3E6")   # 淺藍 - Test Suite
L2_FONT = Font(bold=True, color="1F3864")
TC_FILL = PatternFill("solid", fgColor="F2F2F2")   # 淺灰 - Test Case


def _strip_num_prefix(s: str) -> str:
    """'21.EtherCAT' → 'EtherCAT'，'9.網路配置' → '網路配置'"""
    if not s:
        return s
    parts = s.split(".", 1)
    if len(parts) == 2 and parts[0].strip().isdigit():
        return parts[1].strip()
    return s.strip()


def _find_tc_for_suite(suite_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    嘗試以 suite_name 比對 Test Case 的子項目和功能項（精確比對剝除前綴後的名稱）。
    """
    norm = suite_name.strip().lower()

    def match(s) -> bool:
        if pd.isna(s):
            return False
        return _strip_num_prefix(str(s)).lower() == norm

    mask = df["子項目"].apply(match) | df["功能項"].apply(match)
    return df[mask].copy()


def write_sheet4(ws, records: list[dict]):
    """
    階層頁籤：功能分類 → Test Suite（●）→ Test Case（關鍵字自動對應）
    使用 Excel outline（row grouping）讓 Test Case 行可折疊。
    """
    # 欄位定義
    COLS = ["功能分類", "Test Suite", "功能Tag", "Test Case Name", "Description", "腳本狀態", "State"]
    COL_W = [18, 26, 38, 22, 55, 10, 12]

    for c, (h, w) in enumerate(zip(COLS, COL_W), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    df_tc = pd.read_csv(TC_PATH, encoding="utf-8-sig")

    # 取出「●支援」的 (功能分類, Test Suite) 組合，去重
    supported = [r for r in records if r["納入 Suite"] == "是"]
    # 按功能分類分組，每個功能分類下的 Suite 去重（保留順序）
    from collections import OrderedDict
    grouped: dict[str, list[str]] = OrderedDict()
    suite_tags: dict[tuple, list[str]] = {}  # (l1, suite) → [tags]
    for rec in supported:
        l1 = rec["功能分類"]
        suite = rec["Test Suite（功能項目）"]
        tag = rec["功能Tag"]
        grouped.setdefault(l1, [])
        if suite not in grouped[l1]:
            grouped[l1].append(suite)
        suite_tags.setdefault((l1, suite), []).append(tag)

    r = 2
    total_matched = 0
    total_unmatched = 0

    for l1, suites in grouped.items():
        # ── 功能分類 header row ──
        ws.cell(r, 1, l1).font = L1_FONT
        for c in range(1, len(COLS) + 1):
            ws.cell(r, c).fill = L1_FILL
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for suite in suites:
            tags = suite_tags.get((l1, suite), [])
            tag_str = ", ".join(tags)

            # ── Test Suite row (outline_level=1，可折疊) ──
            ws.cell(r, 1, "")
            ws.cell(r, 2, suite).font = L2_FONT
            ws.cell(r, 3, tag_str)
            for c in range(1, len(COLS) + 1):
                ws.cell(r, c).fill = L2_FILL
                ws.cell(r, c).border = BORDER
                ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=False)
            ws.row_dimensions[r].height = 20
            suite_header_row = r
            r += 1

            # ── Test Case rows (outline_level=2，折疊在 Suite 下) ──
            matched_tc = _find_tc_for_suite(suite, df_tc)
            tc_start = r

            if matched_tc.empty:
                ws.cell(r, 4, "（待補對應）").font = Font(italic=True, color="999999")
                for c in range(1, len(COLS) + 1):
                    ws.cell(r, c).fill = TBC_FILL
                    ws.cell(r, c).border = BORDER
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].height = 18
                r += 1
                total_unmatched += 1
            else:
                for _, tc_row in matched_tc.iterrows():
                    ws.cell(r, 1, "")
                    ws.cell(r, 2, "")
                    ws.cell(r, 3, "")
                    ws.cell(r, 4, tc_row.get("Name", ""))
                    ws.cell(r, 5, tc_row.get("Description", ""))
                    ws.cell(r, 6, tc_row.get("腳本狀態", ""))
                    ws.cell(r, 7, tc_row.get("State", ""))
                    for c in range(1, len(COLS) + 1):
                        ws.cell(r, c).fill = TC_FILL
                        ws.cell(r, c).border = BORDER
                        ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=False)
                    ws.row_dimensions[r].outline_level = 1
                    ws.row_dimensions[r].height = 18
                    r += 1
                total_matched += len(matched_tc)

    ws.sheet_view.showOutlineSymbols = True

    # 統計備注
    ws.cell(r + 1, 1, f"※ 自動比對：找到 {total_matched} 筆 Test Case；{total_unmatched} 個 Suite 待補對應")
    ws.cell(r + 1, 1).font = Font(italic=True, color="888888")

    print(f"  Sheet4：自動比對 {total_matched} 筆 Test Case，{total_unmatched} 個 Suite 待補對應")


def run(model: str = "W3A"):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"TestSuite_{model}.xlsx"

    print(f"  讀取 Function Check List…")
    records = parse_fcl()
    total = len(records)
    supported = sum(1 for r in records if r["納入 Suite"] == "是")
    print(f"  功能Tag 共 {total} 筆，W3A ● 支援：{supported} 筆")

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("1_Test Suite 清單")
    write_sheet1(ws1, records)

    ws2 = wb.create_sheet("2_Mapping Table")
    write_sheet2(ws2, records)

    ws3 = wb.create_sheet("3_Test Case 參考")
    write_sheet3(ws3)

    ws4 = wb.create_sheet("4_Suite & Cases 階層")
    write_sheet4(ws4, records)

    wb.save(out_path)
    print(f"  輸出：{out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", default="W3A", help="機種名稱（預設 W3A）")
    args = parser.parse_args()
    run(model=args.model)
